"""
AI-based sensitive data detection using pluggable providers
"""
from typing import Dict, List, Optional
import json
from pathlib import Path
import re

from src.utils.logger import logger

class SensitiveDataDetector:
    """AI-based sensitive data detection using pluggable providers"""

    def __init__(self, provider=None, model: Optional[str] = None, host: Optional[str] = None):
        from src.ai.factory import ProviderFactory
        self.provider = ProviderFactory.create(provider or 'ollama', model=model, host=host)

    def detect_sensitive_data(self, text_chunk: str, file_name: str = None) -> Dict:
        """
        Detect sensitive data in text chunk

        Args:
            text_chunk: Text to analyze

        Returns:
            Detection results dictionary
        """
        # Prepend file name to text if provided
        if file_name:
            text_for_analysis = f"[FILENAME: {file_name}]\n{text_chunk}"
        else:
            text_for_analysis = text_chunk
        # Use provider; merge with simple pattern heuristics as a booster
        provider_result = self.provider.analyze_text(text_for_analysis)
        heuristics = self._check_sensitive_patterns(text_for_analysis)
        
        # Merge heuristic patterns with provider results
        if heuristics and 'detected_patterns' in provider_result:
            provider_result['detected_patterns'] = list(set(provider_result.get('detected_patterns', []) + heuristics))
            
        # Boost sensitivity score if patterns detected but AI didn't catch them
        if heuristics and provider_result.get('sensitivity_score', 1) < 5:
            provider_result['sensitivity_score'] = min(10, provider_result.get('sensitivity_score', 1) + 2)
            if 'explanation' in provider_result:
                provider_result['explanation'] += f" Score boosted due to detected patterns: {', '.join(heuristics)}"
            else:
                provider_result['explanation'] = f"Patterns detected: {', '.join(heuristics)}"
                
        return provider_result

    def _check_sensitive_patterns(self, text: str) -> List[str]:
        """Check for common sensitive data patterns"""
        patterns = []
        
        # Email addresses
        if re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', text):
            patterns.append('email_address')
            
        # Phone numbers (US format)
        if re.search(r'\b\d{3}[-.]?\d{3}[-.]?\d{4}\b', text):
            patterns.append('phone_number')
            
        # Social Security Numbers
        if re.search(r'\b\d{3}-\d{2}-\d{4}\b', text):
            patterns.append('ssn')
            
        # Credit card numbers (basic pattern)
        if re.search(r'\b\d{4}[- ]?\d{4}[- ]?\d{4}[- ]?\d{4}\b', text):
            patterns.append('credit_card')
            
        return patterns

    def process_chunks_by_file(self, chunks: List[Dict]) -> List[Dict]:
        """
        Process chunks and group results by file
        
        Args:
            chunks: List of chunk data with file information
            
        Returns:
            List of file-level detection results
        """
        # Group chunks by file
        files_data = {}
        for chunk in chunks:
            file_path = chunk.get('file_path', 'unknown')
            file_name = Path(file_path).name if file_path else 'unknown'
            if file_path not in files_data:
                files_data[file_path] = {
                    'file_path': file_path,
                    'file_name': file_name,
                    'chunks': [],
                    'total_text': '',
                    'chunk_count': 0
                }
            chunk['file_name'] = file_name
            files_data[file_path]['chunks'].append(chunk)
            files_data[file_path]['total_text'] += chunk.get('content', '') + ' '
            files_data[file_path]['chunk_count'] += 1
        
        # Process each file
        file_results = []
        for file_path, file_info in files_data.items():
            file_result = self._analyze_file(file_info)
            file_results.append(file_result)
        
        logger.info(f"Processed {len(file_results)} files for sensitive data")
        return file_results

    def _analyze_file(self, file_info: Dict) -> Dict:
        """
        Analyze a single file by aggregating chunk results
        
        Args:
            file_info: File information with chunks
            
        Returns:
            File-level detection result
        """
        file_path = file_info['file_path']
        file_name = file_info.get('file_name', Path(file_path).name if file_path else 'unknown')
        chunks = file_info['chunks']
        total_text = file_info['total_text'].strip()

        # Try to get the original link from the first chunk's metadata
        original_link = 'Unknown'
        if chunks and len(chunks) > 0:
            first_chunk = chunks[0]
            original_link = first_chunk.get('original_link', first_chunk.get('source_url', 'Unknown'))

        # Analyze the entire file text, including file name
        full_file_result = self.detect_sensitive_data(total_text, file_name=file_name)

        # Also analyze individual chunks for detailed breakdown
        chunk_results = []
        max_sensitivity = 1
        all_categories = set()
        all_patterns = set()
        all_explanations = []
        
        for chunk in chunks:
            chunk_result = self.detect_sensitive_data(chunk.get('content', ''))
            chunk_results.append(chunk_result)
            
            # Track maximum sensitivity across chunks
            max_sensitivity = max(max_sensitivity, chunk_result.get('sensitivity_score', 1))
            
            # Collect all categories and patterns
            all_categories.update(chunk_result.get('sensitive_categories', []))
            all_patterns.update(chunk_result.get('detected_patterns', []))
            
            # Collect explanations
            explanation = chunk_result.get('explanation', '')
            if explanation:
                all_explanations.append(explanation)
        
        # Create aggregated file result
        file_result = {
            'file_path': file_path,
            'original_link': original_link,
            'chunk_count': len(chunks),
            'total_text_length': len(total_text),
            'sensitivity_score': max_sensitivity,  # Use highest sensitivity from any chunk
            'avg_sensitivity_score': sum(r.get('sensitivity_score', 1) for r in chunk_results) / len(chunk_results) if chunk_results else 1,
            'confidence': full_file_result.get('confidence', 0.0),
            'sensitive_categories': list(all_categories),
            'detected_patterns': list(all_patterns),
            'explanation': self._create_file_explanation(max_sensitivity, all_categories, all_patterns, all_explanations),
            'recommendations': full_file_result.get('recommendations', []),
            'provider': full_file_result.get('provider', 'ollama'),
            'model': full_file_result.get('model', 'llama3.1'),
            'chunk_details': chunk_results  # Keep individual chunk results for reference
        }
        
        return file_result

    def _create_file_explanation(self, max_sensitivity: int, categories: set, patterns: set, explanations: List[str]) -> str:
        """
        Create a comprehensive explanation for the file-level analysis
        
        Args:
            max_sensitivity: Highest sensitivity score from any chunk
            categories: All sensitive categories detected
            patterns: All patterns detected
            explanations: List of chunk explanations
            
        Returns:
            Comprehensive file explanation
        """
        explanation_parts = []
        
        # Overall sensitivity assessment
        if max_sensitivity >= 8:
            explanation_parts.append(f"This file contains highly sensitive information (score: {max_sensitivity}/10)")
        elif max_sensitivity >= 5:
            explanation_parts.append(f"This file contains moderately sensitive information (score: {max_sensitivity}/10)")
        else:
            explanation_parts.append(f"This file contains low sensitivity information (score: {max_sensitivity}/10)")
        
        # Categories found
        if categories:
            explanation_parts.append(f"Sensitive categories detected: {', '.join(sorted(categories))}")
        
        # Patterns found
        if patterns:
            explanation_parts.append(f"Patterns detected: {', '.join(sorted(patterns))}")
        
        # Key insights from chunk explanations
        if explanations:
            # Take the most relevant explanations (avoid duplicates)
            unique_explanations = list(set(explanations))
            if len(unique_explanations) <= 3:
                explanation_parts.append("Key findings: " + "; ".join(unique_explanations))
            else:
                # Take first few and summarize
                explanation_parts.append(f"Key findings: {unique_explanations[0]}; {len(unique_explanations)-1} additional insights")
        
        return ". ".join(explanation_parts)

    def generate_file_report(self, file_results: List[Dict], output_dir: Path) -> str:
        """
        Generate an Excel report from file-level detection results
        
        Args:
            file_results: List of file-level detection result dictionaries
            output_dir: Directory to save the report
            
        Returns:
            Path to the generated Excel report file
        """
        logger.info(f"Starting Excel report generation for {len(file_results)} files")
        report_path = output_dir / "file_sensitivity_report.xlsx"
        logger.info(f"Report will be saved to: {report_path}")
        
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Create summary data for the main sheet
            summary_data = []
            for result in file_results:
                file_path = result.get('file_path', 'unknown')
                filename = Path(file_path).name if file_path != 'unknown' else 'Unknown'
                
                # Try to get the original link from the file path or metadata
                original_link = result.get('original_link', 'Unknown')
                if original_link == 'Unknown' and 'box_file_' in file_path:
                    # For Box files, we might need to reconstruct or get from metadata
                    original_link = f"Box file: {filename}"
                
                summary_data.append({
                    'Filename': filename,
                    'Original Link': original_link,
                    'Sensitivity Score (Max)': result.get('sensitivity_score', 1),
                    'Sensitivity Score (Avg)': round(result.get('avg_sensitivity_score', 1), 1),
                    'Risk Level': self._get_risk_level(result.get('sensitivity_score', 1)),
                    'Confidence': f"{result.get('confidence', 0.0):.1%}",
                    'Chunks Analyzed': result.get('chunk_count', 0),
                    'Categories': ', '.join(result.get('sensitive_categories', [])),
                    'Patterns': ', '.join(result.get('detected_patterns', [])),
                    'Explanation': result.get('explanation', 'No explanation provided'),
                    'Recommendations': '; '.join(result.get('recommendations', []))
                })
            
            # Create DataFrame and sort by sensitivity score (highest first)
            df = pd.DataFrame(summary_data)
            df = df.sort_values('Sensitivity Score (Max)', ascending=False)
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Sensitivity Analysis"
            
            # Add title
            ws['A1'] = "File-Level Data Sensitivity Analysis Report"
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:J1')
            
            # Add summary statistics
            total_files = len(file_results)
            high_sensitivity_files = sum(1 for r in file_results if r.get('sensitivity_score', 1) >= 7)
            medium_sensitivity_files = sum(1 for r in file_results if 4 <= r.get('sensitivity_score', 1) < 7)
            low_sensitivity_files = sum(1 for r in file_results if r.get('sensitivity_score', 1) < 4)
            avg_sensitivity = sum(r.get('sensitivity_score', 1) for r in file_results) / total_files if total_files > 0 else 0
            
            summary_stats = [
                ['Summary Statistics', ''],
                ['Total Files Analyzed', total_files],
                ['Average Sensitivity Score', f"{avg_sensitivity:.1f}/10"],
                ['High Sensitivity Files (7-10)', f"{high_sensitivity_files} ({high_sensitivity_files/total_files*100:.1f}%)"],
                ['Medium Sensitivity Files (4-6)', f"{medium_sensitivity_files} ({medium_sensitivity_files/total_files*100:.1f}%)"],
                ['Low Sensitivity Files (1-3)', f"{low_sensitivity_files} ({low_sensitivity_files/total_files*100:.1f}%)"],
                ['', ''],
                ['Risk Level Legend', ''],
                ['🔴 HIGH RISK', 'Sensitivity Score 7-10'],
                ['🟡 MEDIUM RISK', 'Sensitivity Score 4-6'],
                ['🟢 LOW RISK', 'Sensitivity Score 1-3'],
                ['', '']
            ]
            
            # Add summary statistics to worksheet
            for i, (label, value) in enumerate(summary_stats, start=3):
                ws[f'A{i}'] = label
                ws[f'B{i}'] = value
                if label and not label.startswith('🔴') and not label.startswith('🟡') and not label.startswith('🟢'):
                    ws[f'A{i}'].font = Font(bold=True)
            
            # Add headers for data
            headers = list(df.columns)
            header_row = len(summary_stats) + 3
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Add data rows
            for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), header_row + 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Color code based on risk level
                    risk_level = row_data[4] if len(row_data) > 4 else ''  # Risk Level column
                    if '🔴' in risk_level:
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    elif '🟡' in risk_level:
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    elif '🟢' in risk_level:
                        cell.fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
                    
                    # Wrap text for explanation and recommendations columns
                    if col_idx in [9, 10]:  # Explanation and Recommendations columns
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = None
                for cell in column:
                    try:
                        # Skip merged cells
                        if hasattr(cell, 'column_letter'):
                            column_letter = cell.column_letter
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                
                if column_letter:
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create detailed analysis sheet
            ws2 = wb.create_sheet("Detailed Analysis")
            
            # Add detailed chunk information
            ws2['A1'] = "Detailed Chunk Analysis"
            ws2['A1'].font = Font(size=16, bold=True)
            ws2.merge_cells('A1:H1')
            
            # Headers for detailed sheet
            detail_headers = ['Original Link', 'Chunk #', 'Chunk Text Length', 'Sensitivity Score', 'Categories', 'Patterns', 'Explanation', 'Recommendations']
            for col, header in enumerate(detail_headers, 1):
                cell = ws2.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Add detailed chunk data
            row_idx = 4
            for file_result in file_results:
                original_link = file_result.get('original_link', 'Unknown')
                chunk_details = file_result.get('chunk_details', [])
                
                for chunk_idx, chunk_result in enumerate(chunk_details, 1):
                    ws2.cell(row=row_idx, column=1, value=original_link)
                    ws2.cell(row=row_idx, column=2, value=chunk_idx)
                    ws2.cell(row=row_idx, column=3, value=len(chunk_result.get('content', '')))
                    ws2.cell(row=row_idx, column=4, value=chunk_result.get('sensitivity_score', 1))
                    ws2.cell(row=row_idx, column=5, value=', '.join(chunk_result.get('sensitive_categories', [])))
                    ws2.cell(row=row_idx, column=6, value=', '.join(chunk_result.get('detected_patterns', [])))
                    ws2.cell(row=row_idx, column=7, value=chunk_result.get('explanation', ''))
                    ws2.cell(row=row_idx, column=8, value='; '.join(chunk_result.get('recommendations', [])))
                    
                    # Color code based on sensitivity score
                    score = chunk_result.get('sensitivity_score', 1)
                    if score >= 7:
                        for col in range(1, 9):
                            ws2.cell(row=row_idx, column=col).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    elif score >= 4:
                        for col in range(1, 9):
                            ws2.cell(row=row_idx, column=col).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    else:
                        for col in range(1, 9):
                            ws2.cell(row=row_idx, column=col).fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
                    
                    row_idx += 1
            
            # Auto-adjust column widths for detailed sheet
            for column in ws2.columns:
                max_length = 0
                column_letter = None
                for cell in column:
                    try:
                        # Skip merged cells
                        if hasattr(cell, 'column_letter'):
                            column_letter = cell.column_letter
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                
                if column_letter:
                    adjusted_width = min(max_length + 2, 50)
                    ws2.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook
            wb.save(report_path)
            logger.info(f"Generated Excel sensitivity report: {report_path}")
            return str(report_path)
            
        except ImportError as e:
            logger.error(f"Excel generation failed - missing dependency: {e}")
            # Fallback to CSV
            csv_path = output_dir / "file_sensitivity_report.csv"
            df = pd.DataFrame(summary_data)
            df = df.sort_values('Sensitivity Score (Max)', ascending=False)
            df.to_csv(csv_path, index=False)
            logger.info(f"Generated CSV sensitivity report: {csv_path}")
            return str(csv_path)
            
        except Exception as e:
            logger.error(f"Excel report generation failed: {e}")
            return ""

    def _get_risk_level(self, sensitivity_score: int) -> str:
        """Get risk level emoji and text based on sensitivity score"""
        if sensitivity_score >= 7:
            return "🔴 HIGH RISK"
        elif sensitivity_score >= 4:
            return "🟡 MEDIUM RISK"
        else:
            return "🟢 LOW RISK" 